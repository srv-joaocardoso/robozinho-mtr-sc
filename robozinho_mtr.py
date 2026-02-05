import sys
from playwright.sync_api import sync_playwright
from dotenv import load_dotenv
import os
import pyodbc
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd  
import io
import time
import re
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, messagebox

def carregar_env():
    if getattr(sys, 'frozen', False):
        # Executável (PyInstaller)
        base_dir = os.path.dirname(sys.executable)
    else:
        # Execução normal (python)
        base_dir = os.path.dirname(os.path.abspath(__file__))

    env_path = os.path.join(base_dir, ".env")
    load_dotenv(env_path)

carregar_env()

def escolher_local_salvar():
    print("Escolha o local para salvar o relatório MTR.")
    root = Tk()
    root.withdraw()

    caminho = filedialog.asksaveasfilename(
        title="Salvar Relatório MTR",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile="relatorio_mtr_balanca.xlsx"
    )

    root.destroy()
    print(f"Caminho selecionado: {caminho}")
    return caminho

arquivo_saida = escolher_local_salvar()
if not arquivo_saida:
    print("Usuário cancelou.")
    sys.exit(0)

def pintar_linha_excel(arquivo, codigo_mtr, erro=False):
    print(f"Pintando linha do MTR {codigo_mtr} como {'Erro' if erro else 'OK'} no Excel.")
    wb = load_workbook(arquivo)
    ws = wb.active

    FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_ERRO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill = FILL_OK if not erro else FILL_ERRO

    col_mtr = None
    col_status = None

    # Localizar colunas
    for idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value).strip()
        if header == "MTR Nº":
            col_mtr = idx
        elif header == "Status":
            col_status = idx

    if not col_mtr:
        print("Coluna 'MTR Nº' não encontrada")
        return

    if not col_status:
        print("Coluna 'Status' não encontrada")
        return

    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=col_mtr).value

        if str(valor) == str(codigo_mtr):
            # Pintar a linha inteira
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

            # Atualizar Status
            if not erro:
                ws.cell(row=row, column=col_status).value = "Enviado"
            else:
                ws.cell(row=row, column=col_status).value = "Erro"

            wb.save(arquivo)
            return

    print(f"MTR {codigo_mtr} não encontrado no Excel")

def conectar_banco():
    print("Conectando ao banco de dados...")
    conn = pyodbc.connect(
        f"DRIVER={{{os.getenv('DB_DRIVER')}}};"
        f"SERVER={os.getenv('DB_HOST')};"
        f"DATABASE={os.getenv('DB_NOME')};"
        f"UID={os.getenv('DB_USUARIO')};"
        f"PWD={os.getenv('DB_SENHA')};"
        f"TrustServerCertificate=yes;"
    )

    return conn

def mascarar_cpf_cnpj(valor):
    if pd.isna(valor):
        return None
    numeros = re.sub(r"\D", "", str(valor))
    if len(numeros) == 11:  # CPF
        return f"{numeros[:3]}.{numeros[3:6]}.{numeros[6:9]}-{numeros[9:]}"
    elif len(numeros) == 14:  # CNPJ
        return f"{numeros[:2]}.{numeros[2:5]}.{numeros[5:8]}/{numeros[8:12]}-{numeros[12:]}"
    else:
        return valor

def login_mtr(page):
    print("Realizando login no MTR...")
    try:
        page.wait_for_selector("#txtCnpj").fill(os.getenv('MTR_LOGIN_CNPJ'))
        page.keyboard.press("Tab")
        page.wait_for_selector("#txtCpfUsuario").fill(os.getenv('MTR_LOGIN_CPF'))
        page.wait_for_selector("#txtSenha").fill(os.getenv('MTR_LOGIN_SENHA'))
        page.wait_for_selector("#btEntrar").click()
        page.wait_for_url("https://mtr.ima.sc.gov.br/ControllerServlet?acao=paginaPrincipal")
    except Exception as e:
        print("Erro no login:", e)

def processar_dados_mtr(page):
    print("Processando dados do MTR...")
    # Gerar XLS do MTR
    data_inicio = (datetime.now() - relativedelta(months=3)).strftime("%d/%m/%Y")
    data_fim = (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")
    url_relatorio = (
        "https://mtr.ima.sc.gov.br/ControllerServlet?"
        "acao=relatorio"
        "&nomeRelatorio=rel_mtr_ger_trans_des"
        "&condicao=pageRelatorio"
        f"&txtDataInicial_Mtr={data_inicio}"
        f"&txtDataFinal_Mtr={data_fim}"
        f"&destinador_codigo={os.getenv('CODIGO_RECEBEDOR_IMA')}"
        "&gerador_codigo=0"
        "&transportador_codigo=0"
    )
    with page.expect_download() as download_info:
        page.evaluate(f"""
            () => {{
                window.location.href = "{url_relatorio}";
            }}
            """
        )  
    download = download_info.value
    caminho_temporario = download.path()
    with open(caminho_temporario, "rb") as f:
        conteudo_bytes = f.read()
    xls = io.BytesIO(conteudo_bytes)
        
    # Aplica Filtro e Extrai Dados
    df_mtr = pd.read_excel(xls)
    df_mtr_salvo = df_mtr[df_mtr["Situação"] == "Salvo"].copy()
    df_mtr_salvo["nota_fiscal"] = (
        df_mtr_salvo["Observações"]
        .astype(str)
        .str.extract(r"(\d+)")
        .astype("Int64")
    )

    # Conectar no Banco e Puxar Dados da Balança
    sql = """
    SELECT
        [Ticket],
        [Veículo],
        [Emissor],
        [Produto],
        [Nota fiscal],
        [Data e Hora de saída],
        [Peso liquido (kg)],
        [Observação]
    FROM Supervisor.dbo.MeioAmbiente
    """
    conn = conectar_banco()
    df_db = pd.read_sql(sql, conn)
    conn.close()

    # Garantir que "Nota fiscal" seja do tipo Int64
    df_db["Nota fiscal"] = df_db["Nota fiscal"].astype("Int64")

    # Merge dos DataFrames
    df_final = df_mtr_salvo.merge(
        df_db,
        left_on="nota_fiscal",
        right_on="Nota fiscal",
        how="left",
        suffixes=("_mtr", "_db")
    )
    
    # Preparar DataFrame para Excel
    df_excel = pd.DataFrame({
        "Balança Data e Hora Pesagem": df_final["Data e Hora de saída"],
        "Nota fiscal": df_final["nota_fiscal"],
        "MTR Nº": df_final["MTR Nº"],
        "Balança Placa": df_final["Veículo"],
        "Correção MTR": "",
        "MTR Residuo Código": df_final["Residuo código/descrição"].astype(str).str[:6],
        "MTR Classe": df_final["Classe"],
        "Balança Qt. Tonelada": df_final["Peso liquido (kg)"] / 1000,
        "MTR Gerador Nome": df_final["Gerador Nome"],
        "MTR Gerador CPF/CNPJ": df_final["Gerador CPF/CNPJ"],
        "MTR Transportador Nome": df_final["Transportador Nome"],
        "MTR Transportador CPF/CNPJ": df_final["Transportador CPF/CNPJ"],
        "MTR Motorista": df_final["Motorista"],
        "MTR Placa": df_final["Placa"],
        "MTR Data de Emissão": df_final["Data de Emissão"],
        "MTR Qt. Tonelada": df_final["Qt. tonelada"],
        "MTR Observações": df_final["Observações"],
        "MTR Tecnologia": df_final["Tecnologia"],
        "Balança Ticket": df_final["Ticket"],
        "Balança Emissor": df_final["Emissor"],
        "Balança Produto": df_final["Produto"],
        "Balança Observação": df_final["Observação"]
    })

    # Aplica Formatações
    df_excel["MTR Gerador CPF/CNPJ"] = (
        df_excel["MTR Gerador CPF/CNPJ"]
        .astype(str)
        .apply(mascarar_cpf_cnpj)
    )
    df_excel["MTR Data de Emissão"] = pd.to_datetime(
        df_excel["MTR Data de Emissão"],
        errors="coerce"
    )

    # Criar coluna "Status" e a define como primeira
    df_excel["Status"] = df_excel["Balança Data e Hora Pesagem"].apply(
        lambda x: "Pendente" if pd.isna(x) else "Aberto"
    )
    colunas = ["Status"] + [c for c in df_excel.columns if c != "Status"]
    df_excel = df_excel[colunas]

    # Ordenar o DataFrame
    df_excel.sort_values(
        by=["Balança Data e Hora Pesagem", "MTR Data de Emissão"],
        inplace=True
    )

    # Salvar DataFrame em Excel
    df_excel.to_excel(
        arquivo_saida,
        index=False,
        engine="openpyxl"
    )

    # Abrir o arquivo salvo para aplicar formatações adicionais
    wb = load_workbook(arquivo_saida)
    ws = wb.active

    # Adicionar Tabela
    ultima_coluna = get_column_letter(ws.max_column)
    ultima_linha = ws.max_row
    tabela = Table(
        displayName="TabelaMTR",
        ref=f"A1:{ultima_coluna}{ultima_linha}"
    )

    # Definir estilo da tabela
    style = TableStyleInfo(
        name="TableStyleMedium3",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabela.tableStyleInfo = style
    ws.add_table(tabela)

    # Aplicar preenchimento amarelo para células que não atendem aos critérios
    fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    col_residuo = df_excel.columns.get_loc("MTR Residuo Código") + 1
    col_classe = df_excel.columns.get_loc("MTR Classe") + 1
    col_correcao = df_excel.columns.get_loc("Correção MTR") + 1
    col_data = df_excel.columns.get_loc("MTR Data de Emissão") + 1
    for row in range(2, ws.max_row + 1):
        correcoes = []
        if ws.cell(row=row, column=col_residuo).value != os.getenv('CODIGO_RESIDUO_PADRAO'):
            ws.cell(row=row, column=col_residuo).fill = fill_amarelo
            correcoes.append(f"Alterado para {os.getenv('CODIGO_RESIDUO_PADRAO')} (IN 13/2012-IBAMA)")
        if ws.cell(row=row, column=col_classe).value != os.getenv('CLASSE_PADRAO'):
            ws.cell(row=row, column=col_classe).fill = fill_amarelo
            correcoes.append(f"Alterado para classe {os.getenv('CLASSE_PADRAO')} (NBR 10.004)")
        if correcoes:
            correcoes = ["Quantidade corrigida de acordo com o peso líquido de entrada"] + correcoes
            ws.cell(row=row, column=col_correcao).value = ",\n".join(correcoes) + "."
        ws.cell(row=row, column=col_data).number_format = "DD/MM/YYYY"

    # Ocultar colunas desnecessárias
    colunas_ocultar = [
        "MTR Gerador CPF/CNPJ",
        "MTR Transportador Nome",
        "MTR Transportador CPF/CNPJ",
        "MTR Motorista",
        "MTR Placa",
        "MTR Data de Emissão",
        "MTR Qt. Tonelada",
        "MTR Observações",
        "MTR Tecnologia",
        "Balança Ticket",
        "Balança Emissor",
        "Balança Produto",
        "Balança Observação"
    ]
    for idx, nome in enumerate(df_excel.columns, start=1):
        if nome in colunas_ocultar:
            ws.column_dimensions[get_column_letter(idx)].hidden = True
    for col in ws.columns:
        col_letter = col[0].column_letter
        if not ws.column_dimensions[col_letter].hidden:
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col_letter].width = max_length + 2

    # Salvar alterações no arquivo Excel
    wb.save(arquivo_saida)
    return arquivo_saida

def ir_para_manifesto(page):
    pagina_manifesto = "https://mtr.ima.sc.gov.br/ControllerServlet?acao=acompanhamentoManifesto"
    if page.url != pagina_manifesto:
        page.evaluate(f"""
            () => {{
                window.location.href = "{pagina_manifesto}";
            }}
            """
        )
    page.wait_for_url(pagina_manifesto)

def receber_mtr(page, codigo_mtr, data_recebimento, peso, correcao = None):
    print(f"Recebendo MTR {codigo_mtr} - Peso: {peso} - Data: {data_recebimento} - Correção: {correcao}")

    # Navegar até a página de acompanhamento do MTR
    ir_para_manifesto(page)
    
    # Abre Menu Recebimento de MTR
    print("Abrindo menu de recebimento de MTR...")
    page.wait_for_selector("#txtCodigoMtrRecebimento")
    page.wait_for_selector("#txtCodigoMtrRecebimento").fill(codigo_mtr)
    page.wait_for_selector("#btnReceberMtr").click()

    # Aguardar carregamento do modal de recebimento
    print("Aguardando modal de recebimento...")
    modalRecebimento = page.locator("#divRecebimento")
    modalRecebimento.wait_for()

    # Interagir com o calendário para selecionar a data de recebimento
    print("Selecionando data de recebimento...")
    page.wait_for_selector("#formRespRecebimento > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > img").click()
    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    mes_calendario = page.locator("#ui-datepicker-div > div > div > span.ui-datepicker-month")
    ano_calendario = page.locator("#ui-datepicker-div > div > div > span.ui-datepicker-year")
    dia_recebimento = str(int(data_recebimento.split("/")[0]))
    mes_recebimento = meses[int(data_recebimento.split("/")[1]) - 1]
    ano_recebimento = data_recebimento.split("/")[2]
    seletor_voltar_mes = page.locator("#ui-datepicker-div > div > a.ui-datepicker-prev.ui-corner-all > span")
    seletor_voltar_mes.wait_for()
    while mes_calendario.inner_text() != mes_recebimento or ano_calendario.inner_text() != ano_recebimento:
        seletor_voltar_mes.click()
    tabela = page.locator("table.ui-datepicker-calendar")
    tabela.wait_for()
    tabela.get_by_role("link", name=dia_recebimento, exact=True).click()
    
    #Selecionar recebedor
    print("Selecionando recebedor...")
    page.evaluate("pesquisaResponsavelRecebimento()")
    tabela = page.locator("#pesquisaResponsavelRecebimento")
    tabela.wait_for()
    tabela.locator(
        "td",
        has_text=os.getenv("RECEBEDOR")
    ).first.click()
    linha = page.locator("#tbRecebeMTR tbody tr").first

    def jutificar_correcao(correcao_texto):
        # Abre e define menu de correção
        print("Definindo correção de recebimento...")
        linha.locator("td").nth(5).locator("img").click()
        page.wait_for_selector("#txtJust").fill(correcao_texto)
        modalTxtJust = page.locator("div.ui-dialog:has(#txtJust):visible").first
        modalTxtJust.locator("div.ui-dialog-buttonpane button:has-text('Salvar')").click()

    if pd.isna(correcao):
        # Define peso de recebimento
        print("Definindo peso de recebimento...")
        linha.locator("td").nth(4).locator("input").fill((f"{peso:.5f}").replace(".", ","))
        peso_mtr = float(linha.locator("td").nth(3).inner_text().replace("(Ton)", "").replace(",", ".").strip())
        if peso_mtr != peso:
            print(peso_mtr, peso)
            jutificar_correcao("Quantidade corrigida de acordo com o peso líquido de entrada.")

    else:
        jutificar_correcao(correcao)
        
        # Abre menu de justificativa
        print("Abrindo menu de justificativa...")
        linha.locator("td").nth(6).locator("input").click()

        #Adiciona resíduo
        print("Adicionando resíduo...")
        page.wait_for_selector("#linkResiduo").click()
        page.wait_for_selector("#txtTipoResiduo").fill(os.getenv("CODIGO_RESIDUO_PADRAO"))
        page.wait_for_selector("#txtQuantidade").fill(f"{peso:.2f}")
        time.sleep(0.2)
        page.select_option("#cmbUnidade", label="Tonelada")
        time.sleep(0.2)
        page.select_option("#cmbEstadoFisico", label="Sólido")
        time.sleep(0.2)
        page.select_option("#cmbClasse", label=os.getenv("CLASSE_PADRAO"))
        time.sleep(0.2)
        page.select_option("#cmbAcondicionamento", label="E03 - Caçamba Fechada")
        time.sleep(0.2)
        page.select_option("#cmbTecnologia", label="Recuperação energética")
        page.wait_for_selector("body > div:nth-child(12) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1)").click()
        page.locator("div.ui-dialog:has-text('Resíduo Cadastrado Com Sucesso!'):visible button[title='close']").click()

    if os.getenv("PRODUCAO") == "1":
        print("Confirmando recebimento...")
        page.evaluate("salvaReceber()")
        page.locator(
            "div.ui-dialog:has(#msg):visible button.ui-dialog-titlebar-close"
        ).click()
    else:
        print("Ambiente de teste - não confirmando recebimento.")
        
        page.screenshot(path=f"sucesso_mtr_{codigo_mtr}.png")
        modalRecebimento.nth(-1).locator("xpath=../div[contains(@class,'ui-dialog-titlebar')]//button[@title='close']").click()

def processar_correcoes_mtr(page, arquivo):
    df = pd.read_excel(arquivo)

    # Filtrar apenas MTRs em aberto
    df = df[df["Status"].str.strip().str.lower() == "aberto"]

    # Garante tipos corretos
    df["MTR Nº"] = df["MTR Nº"].astype(str)
    df["Balança Qt. Tonelada"] = pd.to_numeric(df["Balança Qt. Tonelada"], errors="coerce")

    for _, row in df.iterrows():
        codigo_mtr = row["MTR Nº"]
        peso = row["Balança Qt. Tonelada"]
        data_recebimento = row["Balança Data e Hora Pesagem"].strftime('%d/%m/%Y') 
        correcao = row["Correção MTR"]

        # Data no formato esperado pelo calendário
        try:
            receber_mtr(
                page=page,
                codigo_mtr=codigo_mtr,
                data_recebimento=data_recebimento,
                peso=peso,
                correcao=correcao
            )
            pintar_linha_excel(arquivo, codigo_mtr)
        except Exception as e:
            page.screenshot(path=f"erro_mtr_{codigo_mtr}.png")
            print(f"Erro ao processar MTR {codigo_mtr}: {e}")
            pintar_linha_excel(arquivo, codigo_mtr, erro=True)
            ir_para_manifesto(page)


with sync_playwright() as p:
    # Abrir Navegador e Navegar até a página de login do MTR
    navegador = p.chromium.launch(headless=os.getenv("HEADLESS") == "1")
    contexto = navegador.new_context(accept_downloads=True)
    pagina = contexto.new_page()
    pagina.goto("https://mtr.ima.sc.gov.br/")

    login_mtr(pagina)
    arquivo = processar_dados_mtr(pagina)
    processar_correcoes_mtr(pagina, arquivo)
    
    navegador.close()